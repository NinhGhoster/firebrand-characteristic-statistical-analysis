import openpyxl
import os

file_path = "./Bulk density/bulk density & MC data.xlsx"

print(f"Inspecting formulas in {file_path}...")
try:
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        
        # Check a few rows for formulas
        formulas_found = 0
        for row in ws.iter_rows(min_row=2, max_row=10):
            for cell in row:
                if cell.data_type == 'f': # 'f' indicates a formula
                    print(f"  Found formula in {cell.coordinate}: {cell.value}")
                    formulas_found += 1
            if formulas_found >= 3: # Limit output
                break
        
        if formulas_found == 0:
            print("  No formulas found in first 10 rows.")

except Exception as e:
    print(f"Error inspecting {file_path}: {e}")
