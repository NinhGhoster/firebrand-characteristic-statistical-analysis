import openpyxl
from openpyxl.styles import Font, PatternFill
from copy import copy

# Load the template file (stringybark)
template_wb = openpyxl.load_workbook('/Users/firecaster/Library/CloudStorage/OneDrive-TheUniversityofMelbourne/Documents/Chapter 2/Chapter 2 data/Stringybark/stringybark.xlsx')
template_ws = template_wb.active

# Get the header formatting from template
template_header_cell = template_ws['A1']
header_font = copy(template_header_cell.font)
header_fill = copy(template_header_cell.fill)

print("Template formatting detected:")
print(f"Header font bold: {header_font.b}")
print(f"Header fill pattern: {header_fill.patternType}")
print(f"Template has autofilter: {template_ws.auto_filter.ref if template_ws.auto_filter else 'No'}")

# Load the Combined_Data file
data_wb = openpyxl.load_workbook('/Users/firecaster/Library/CloudStorage/OneDrive-TheUniversityofMelbourne/Documents/Chapter 2/Chapter 2 data/Candle bark/Combined_Data.xlsx')

# Apply formatting to each sheet
for sheet_name in data_wb.sheetnames:
    ws = data_wb[sheet_name]
    print(f"\nFormatting sheet: {sheet_name}")
    
    # Format header row (row 1)
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value is not None:
                cell.font = copy(header_font)
                cell.fill = copy(header_fill)
    
    # Add autofilter to header row
    if ws.max_row > 0:
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}1"
        print(f"  Autofilter added to: A1:{openpyxl.utils.get_column_letter(ws.max_column)}1")

# Save the formatted file
data_wb.save('/Users/firecaster/Library/CloudStorage/OneDrive-TheUniversityofMelbourne/Documents/Chapter 2/Chapter 2 data/Candle bark/Combined_Data.xlsx')
print("\n✓ Formatting applied and file saved!")
