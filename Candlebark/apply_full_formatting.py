import openpyxl
from openpyxl.styles import Font, PatternFill
from copy import copy

# Load template
template_wb = openpyxl.load_workbook('/Users/firecaster/Library/CloudStorage/OneDrive-TheUniversityofMelbourne/Documents/Chapter 2/Chapter 2 data/Stringybark/stringybark.xlsx')
template_ws = template_wb.active

# Get all properties from template header
template_cell = template_ws['A1']
print("Stringybark template header cell properties:")
print(f"Value: {template_cell.value}")
print(f"Font: {template_cell.font}")
print(f"Font bold: {template_cell.font.b}")
print(f"Font size: {template_cell.font.sz}")
print(f"Font name: {template_cell.font.name}")
print(f"Font color object: {template_cell.font.color}")
print(f"Fill: {template_cell.fill}")
print(f"Fill pattern: {template_cell.fill.patternType}")
print(f"Number format: {template_cell.number_format}")
print(f"Alignment: {template_cell.alignment}")
print(f"Border: {template_cell.border}")

# Now copy all formatting to Combined_Data
data_wb = openpyxl.load_workbook('/Users/firecaster/Library/CloudStorage/OneDrive-TheUniversityofMelbourne/Documents/Chapter 2/Chapter 2 data/Candle bark/Combined_Data.xlsx')

for sheet_name in data_wb.sheetnames:
    ws = data_wb[sheet_name]
    
    # Apply full formatting to header
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value is not None:
                # Copy all formatting
                cell.font = copy(template_cell.font)
                cell.fill = copy(template_cell.fill)
                cell.number_format = template_cell.number_format
                cell.border = copy(template_cell.border)
                cell.alignment = copy(template_cell.alignment)

data_wb.save('/Users/firecaster/Library/CloudStorage/OneDrive-TheUniversityofMelbourne/Documents/Chapter 2/Chapter 2 data/Candle bark/Combined_Data.xlsx')
print("\n✓ Complete formatting applied!")
