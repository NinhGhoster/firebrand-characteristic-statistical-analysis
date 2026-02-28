import openpyxl
import pandas as pd

# Load Combined_Data file
excel_file = '/Users/firecaster/Library/CloudStorage/OneDrive-TheUniversityofMelbourne/Documents/Chapter 2/Chapter 2 data/Candle bark/Combined_Data.xlsx'
xls = pd.ExcelFile(excel_file)

# Create a new workbook with standardized column names
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Load template formatting
template_wb = openpyxl.load_workbook('/Users/firecaster/Library/CloudStorage/OneDrive-TheUniversityofMelbourne/Documents/Chapter 2/Chapter 2 data/Stringybark/stringybark.xlsx')
template_ws = template_wb.active
template_cell = template_ws['A1']

from copy import copy

for sheet_name in xls.sheet_names:
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    
    # Standardize column names - convert mm³/mm² to mm3/mm2
    df.columns = df.columns.str.replace('mm³', 'mm3').str.replace('mm²', 'mm2')
    
    # Create new sheet
    ws = wb.create_sheet(sheet_name)
    
    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        # Apply formatting
        cell.font = copy(template_cell.font)
        cell.fill = copy(template_cell.fill)
        cell.border = copy(template_cell.border)
        cell.alignment = copy(template_cell.alignment)
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx).value = value
    
    # Add autofilter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}1"
    
    print(f"✓ {sheet_name}: {len(df)} rows")

# Save the file
wb.save(excel_file)
print(f"\n✓ Column names standardized to mm3/mm2")
print(f"✓ File saved: {excel_file}")
