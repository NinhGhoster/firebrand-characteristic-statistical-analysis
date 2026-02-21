import openpyxl
import os
import re

xlsx_files = [
    "./Bulk density/bulk density & MC data.xlsx",
    "./Stringybark/stringybark.xlsx",
    "./Candle bark/Combined_Data.xlsx",
    "./Branchlet/branchlet raw data.xlsx",
    "./Branchlet/branchlet raw data_pre_cali.xlsx"
]

files_updated = 0

for file_path in xlsx_files:
    if not os.path.exists(file_path):
        continue
    
    print(f"Checking {file_path}...")
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        workbook_modified = False
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Assuming headers are in the first row
            # If sheet is empty, max_row might be 1 but values None.
            if ws.max_row >= 1:
                header_row = ws[1]
                sheet_modified = False
                
                for cell in header_row:
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        new_value = original_value
                        
                        # Apply regex replacement
                        if re.search(r'Volume.*mm', original_value, re.IGNORECASE) and 'mm3' not in original_value:
                             new_value = 'Volume (mm3)'
                        elif re.search(r'Surface Area.*mm', original_value, re.IGNORECASE) and 'mm2' not in original_value:
                             new_value = 'Surface Area (mm2)'
                        
                        if new_value != original_value:
                            print(f"  Updated header in '{sheet_name}': '{original_value}' -> '{new_value}'")
                            cell.value = new_value
                            sheet_modified = True
                
                if sheet_modified:
                    workbook_modified = True
        
        if workbook_modified:
            wb.save(file_path)
            print(f"Saved changes to {file_path}\n")
            files_updated += 1
        else:
            print(f"No changes needed for {file_path}\n")
            
    except Exception as e:
        print(f"Error processing {file_path}: {e}\n")

print(f"Total Excel files updated: {files_updated}")
